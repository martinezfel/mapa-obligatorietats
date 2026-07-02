import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

with open("serveis.json", encoding="utf-8") as f:
    serveis = json.load(f)

wb = Workbook()

HEADER_FILL = PatternFill("solid", start_color="935580", end_color="935580")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

# ---- Sheet 1: Serveis (un per fila) ----
ws1 = wb.active
ws1.title = "Serveis"
cols1 = ["id", "nom", "bloc", "ambit", "tram_base", "font_star", "estat", "resum", "competencia", "data_actualitzacio"]
widths1 = [26, 34, 16, 30, 12, 10, 10, 55, 13, 16]
ws1.append(cols1)
for sv in serveis:
    ws1.append([
        sv["id"], sv["nom"], sv["bloc"], sv["ambit"], sv["tram_base"], sv["font_star"],
        sv["estat"], sv["resum"], sv["competencia"], sv["data_actualitzacio"],
    ])
for i, w in enumerate(widths1, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w
for row in ws1.iter_rows(min_row=2):
    for cell in row:
        cell.font = BODY_FONT
        cell.alignment = WRAP
style_header(ws1, len(cols1))

# ---- Sheet 2: Checklist (un per punt de verificació) ----
ws2 = wb.create_sheet("Checklist")
cols2 = ["servei_id", "ordre", "pregunta", "llei_nom", "article",
         "explicacio", "llindar_label", "pob_min", "pob_min_strict", "pob_max", "pob_max_strict"]
widths2 = [22, 8, 45, 45, 12, 50, 20, 10, 12, 10, 12]
ws2.append(cols2)
for sv in serveis:
    for idx, it in enumerate(sv.get("checklist", []), start=1):
        ws2.append([
            sv["id"], idx, it["pregunta"], it["llei_nom"], it["article"],
            it["explicacio"], it["llindar_label"],
            it["pob_min"], it["pob_min_strict"], it["pob_max"], it["pob_max_strict"],
        ])
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
for row in ws2.iter_rows(min_row=2):
    for cell in row:
        cell.font = BODY_FONT
        cell.alignment = WRAP
style_header(ws2, len(cols2))

# ---- Sheet 3: NormativaMare (lleis de base que apliquen encara que cap pregunta les citi directament) ----
ws_mare = wb.create_sheet("NormativaMare")
cols_mare = ["servei_id", "llei_nom"]
ws_mare.append(cols_mare)
for sv in serveis:
    for llei in sv.get("normativa_mare", []):
        ws_mare.append([sv["id"], llei])
ws_mare.column_dimensions["A"].width = 22
ws_mare.column_dimensions["B"].width = 70
for row in ws_mare.iter_rows(min_row=2):
    for cell in row:
        cell.font = BODY_FONT
        cell.alignment = WRAP
style_header(ws_mare, len(cols_mare))

# ---- Sheet 3: Instruccions ----
ws3 = wb.create_sheet("Instruccions", 0)
instr = [
    ("Com editar aquest fitxer", ""),
    ("", ""),
    ("Pestanya 'Serveis'", "Un fila per servei. Edita 'resum', 'competencia' i 'data_actualitzacio' lliurement."),
    ("", "Per marcar una fitxa com a validada, posa 'estat' = validat (i emplena la pestanya Checklist)."),
    ("", "'tram_base': deixa-ho buit si el servei no té llindar de població (per exemple, tot el bloc d'interès municipal)."),
    ("", "'font_star': VERTADER si el servei mínim només surt al Decret legislatiu 2/2003 (norma catalana), no a la LBRL estatal."),
    ("", ""),
    ("Pestanya 'Checklist'", "Una fila per cada pregunta de verificació d'un servei. 'servei_id' ha de coincidir"),
    ("", "exactament amb l'id del servei a la pestanya 'Serveis'."),
    ("", "La citació que es veu a la pàgina es genera automàticament amb 'llei_nom' + 'article' — no cal"),
    ("", "escriure-la enlloc més ni repetir-la."),
    ("", "'pob_min' / 'pob_max': deixa-ho buit si no aplica. 'pob_min_strict' = VERTADER vol dir "),
    ("", "'més de X' (estricte); FALS vol dir 'X o més' (inclusiu). Igual per pob_max amb 'menys de' / 'com a màxim'."),
    ("", "'llindar_label': el text que agrupa les preguntes quan es consulta sense triar municipi"),
    ("", "(per exemple: 'Tots els municipis', 'Menys de 5.000 hab.', 'Entre 3.000 i 5.000 hab.')."),
    ("", ""),
    ("Pestanya 'NormativaMare'", "Lleis de base que apliquen a tot el servei (per exemple, la LBRL o el Decret"),
    ("", "legislatiu 2/2003) encara que cap pregunta del checklist les citi directament. Apareixeran igualment"),
    ("", "dins de 'Mostra la normativa completa'. Una fila per cada (servei_id, llei_nom)."),
    ("", ""),
    ("Un cop editat", "Puja aquest fitxer al xat i digues que el converteixi de nou a serveis.json."),
]
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 90
for r, (a, b) in enumerate(instr, start=1):
    ws3.cell(row=r, column=1, value=a).font = Font(bold=True, name="Arial", size=10)
    ws3.cell(row=r, column=2, value=b).font = Font(name="Arial", size=10)
    ws3.cell(row=r, column=2).alignment = WRAP

wb.save("serveis_edicio.xlsx")
print("serveis_edicio.xlsx generat")
