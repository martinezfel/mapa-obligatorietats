import json
import math
from openpyxl import load_workbook

def clean(v):
    if v is None:
        return None
    if isinstance(v, str) and v.strip() == "":
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    return v

def to_bool(v, default=True):
    if v is None or v == "":
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in ("vertader", "true", "1", "si", "sí", "yes", "v")

def to_int(v):
    v = clean(v)
    return int(v) if v is not None else None

wb = load_workbook("serveis_edicio.xlsx", data_only=True)
ws1 = wb["Serveis"]
ws2 = wb["Checklist"]

serveis_by_id = {}
order = []
headers1 = [c.value for c in ws1[1]]
for row in ws1.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers1, row))
    if not d.get("id"):
        continue
    sid = str(d["id"]).strip()
    serveis_by_id[sid] = {
        "id": sid,
        "nom": clean(d.get("nom")),
        "bloc": clean(d.get("bloc")),
        "ambit": clean(d.get("ambit")),
        "tram_base": to_int(d.get("tram_base")),
        "font_star": to_bool(d.get("font_star"), default=False),
        "estat": clean(d.get("estat")) or "pendent",
        "resum": clean(d.get("resum")),
        "competencia": clean(d.get("competencia")),
        "normativa_mare": [],
        "checklist": [],
        "data_actualitzacio": clean(d.get("data_actualitzacio")),
    }
    order.append(sid)

if "NormativaMare" in wb.sheetnames:
    ws_mare = wb["NormativaMare"]
    headers_mare = [c.value for c in ws_mare[1]]
    for row in ws_mare.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers_mare, row))
        sid = clean(d.get("servei_id"))
        llei = clean(d.get("llei_nom"))
        if not sid or not llei:
            continue
        sid = str(sid).strip()
        if sid not in serveis_by_id:
            print("AVÍS: servei_id desconegut a NormativaMare, s'ignora la fila:", sid)
            continue
        serveis_by_id[sid]["normativa_mare"].append(llei)

headers2 = [c.value for c in ws2[1]]
rows2 = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers2, row))
    if not d.get("servei_id"):
        continue
    rows2.append(d)
rows2.sort(key=lambda d: (str(d["servei_id"]), to_int(d.get("ordre")) or 0))

for d in rows2:
    sid = str(d["servei_id"]).strip()
    if sid not in serveis_by_id:
        print("AVÍS: servei_id desconegut a Checklist, s'ignora la fila:", sid)
        continue
    serveis_by_id[sid]["checklist"].append({
        "pregunta": clean(d.get("pregunta")),
        "llei_nom": clean(d.get("llei_nom")),
        "article": clean(d.get("article")) or "",
        "explicacio": clean(d.get("explicacio")),
        "llindar_label": clean(d.get("llindar_label")),
        "pob_min": to_int(d.get("pob_min")),
        "pob_min_strict": to_bool(d.get("pob_min_strict"), default=True),
        "pob_max": to_int(d.get("pob_max")),
        "pob_max_strict": to_bool(d.get("pob_max_strict"), default=True),
    })

serveis = [serveis_by_id[sid] for sid in order]

with open("serveis.json", "w", encoding="utf-8") as f:
    json.dump(serveis, f, ensure_ascii=False)

n_validats = sum(1 for x in serveis if x["estat"] == "validat")
print(len(serveis), "serveis importats,", n_validats, "validats")
